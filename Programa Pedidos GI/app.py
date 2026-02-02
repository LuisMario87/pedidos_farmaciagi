import json
import sqlite3
import tkinter as tk
import os
import smtplib
import threading

from tkinter import messagebox, ttk
from openpyxl import Workbook
from datetime import datetime
from email.message import EmailMessage

# =============================
# CONFIGURACIÓN
# =============================
with open("config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

# =============================
# BASE DE DATOS
# =============================
def inicializar_bd():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS medicamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            presentacion TEXT NOT NULL,
            activo INTEGER DEFAULT 1
        )
                   

        
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            folio TEXT,
            fecha TEXT,
            ruta_excel TEXT,
            estado TEXT,
            error TEXT
        )
        """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS pedido_detalle (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pedido_id INTEGER,
        medicamento TEXT,
        cantidad INTEGER
    )
    """)


    conn.commit()
    conn.close()

def obtener_medicamentos(filtro=""):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if filtro:
        cursor.execute("""
            SELECT nombre, presentacion
            FROM medicamentos
            WHERE activo = 1 AND nombre LIKE ?
            ORDER BY nombre
        """, (f"%{filtro}%",))
    else:
        cursor.execute("""
            SELECT nombre, presentacion
            FROM medicamentos
            WHERE activo = 1
            ORDER BY nombre
        """)

    datos = cursor.fetchall()
    conn.close()
    return datos

def guardar_pedido_bd(folio, ruta, estado, pedido):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO pedidos (folio, fecha, ruta_excel, estado, error)
        VALUES (?, ?, ?, ?, NULL)
    """, (folio, datetime.now().strftime("%Y-%m-%d %H:%M"), ruta, estado))

    pedido_id = cursor.lastrowid

    for item in pedido:
        cursor.execute("""
            INSERT INTO pedido_detalle (pedido_id, medicamento, cantidad)
            VALUES (?, ?, ?)
        """, (pedido_id, item["nombre"], item["cantidad"]))

    conn.commit()
    conn.close()

    return pedido_id

def actualizar_estado_bd(folio, estado, error=None):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pedidos
        SET estado = ?, error = ?
        WHERE folio = ?
    """, (estado, error, folio))

    conn.commit()
    conn.close()



# =============================
# EXCEL
# =============================
def generar_folio():
    fecha = datetime.now().strftime("%Y%m%d")
    os.makedirs("pedidos_excel", exist_ok=True)
    contador = len(os.listdir("pedidos_excel")) + 1
    return f"{config['farmacia_id']}-{fecha}-{contador:04d}"

def exportar_pedido_excel(pedido):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    folio = generar_folio()
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws["A1"] = "Farmacia"
    ws["B1"] = config["farmacia_nombre"]
    ws["A2"] = "Folio"
    ws["B2"] = folio
    ws["A3"] = "Fecha"
    ws["B3"] = fecha

    ws["A5"] = "Medicamento"
    ws["B5"] = "Piezas"
    ws["C5"] = "Surtido"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20

    fila = 6
    for item in pedido:
        ws[f"A{fila}"] = item["nombre"]
        ws[f"B{fila}"] = item["cantidad"]
        fila += 1

    ruta = f"pedidos_excel/Pedido_{folio}.xlsx"
    wb.save(ruta)
    return ruta, folio

# =============================
# CORREO
# =============================
def enviar_correo_pedido(ruta_excel, folio):
    msg = EmailMessage()
    msg["Subject"] = f"Pedido {folio} - {config['farmacia_nombre']}"
    msg["From"] = config["correo_emisor"]
    msg["To"] = config["correo_destino"]

    msg.set_content(f"""
Nuevo pedido generado

Farmacia: {config['farmacia_nombre']}
Folio: {folio}
Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}
""")

    with open(ruta_excel, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(ruta_excel)
        )

    with smtplib.SMTP(config["smtp_server"], config["smtp_port"]) as server:
        server.starttls()
        server.login(config["correo_emisor"], config["correo_password"])
        server.send_message(msg)

def enviar_pedido_async(pedido_info):
    pedido_info["estado"] = "Enviando"
    actualizar_estado_bd(pedido_info["folio"], "Enviando")

    try:
        enviar_correo_pedido(pedido_info["ruta"], pedido_info["folio"])
        pedido_info["estado"] = "Enviado"
        pedido_info["error"] = None
        actualizar_estado_bd(pedido_info["folio"], "Enviado")

    except Exception as e:
        pedido_info["estado"] = "Error"
        pedido_info["error"] = str(e)
        actualizar_estado_bd(pedido_info["folio"], "Error", str(e))

# =============================
# INTERFAZ
# =============================
inicializar_bd()

pedido = []
pedidos_pendientes = []

root = tk.Tk()
root.title("Sistema de Pedidos")
root.geometry("750x900")

tk.Label(root, text=f"{config['farmacia_nombre']} - SISTEMA DE PEDIDOS",
         font=("Arial", 14, "bold")).pack(pady=10)

# =============================
# BUSCADOR
# =============================
frame_busqueda = tk.Frame(root)
frame_busqueda.pack(fill="x", padx=10)

tk.Label(frame_busqueda, text="Buscar medicamento:").pack(anchor="w")
entry_buscar = tk.Entry(frame_busqueda)
entry_buscar.pack(fill="x")

tree_med = ttk.Treeview(root, columns=("nombre", "presentacion"),
                        show="headings", height=8)
tree_med.heading("nombre", text="Nombre")
tree_med.heading("presentacion", text="Presentación")
tree_med.column("nombre", width=400)
tree_med.column("presentacion", width=150)
tree_med.pack(fill="both", padx=10, pady=5)

def actualizar_tabla_medicamentos():
    tree_med.delete(*tree_med.get_children())
    for m in obtener_medicamentos(entry_buscar.get()):
        tree_med.insert("", tk.END, values=m)

entry_buscar.bind("<KeyRelease>", lambda e: actualizar_tabla_medicamentos())
actualizar_tabla_medicamentos()

# =============================
# PEDIDO ACTUAL
# =============================
tk.Label(root, text="Pedido actual", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)

tree_pedido = ttk.Treeview(root, columns=("nombre", "cantidad"),
                           show="headings", height=6)
tree_pedido.heading("nombre", text="Nombre")
tree_pedido.heading("cantidad", text="Piezas")
tree_pedido.pack(fill="x", padx=10)

def actualizar_tabla_pedido():
    tree_pedido.delete(*tree_pedido.get_children())
    for p in pedido:
        tree_pedido.insert("", tk.END, values=(p["nombre"], p["cantidad"]))

# =============================
# AGREGAR
# =============================
frame_cantidad = tk.Frame(root)
frame_cantidad.pack(pady=5)

tk.Label(frame_cantidad, text="Cantidad:").pack(side="left")
entry_cantidad = tk.Entry(frame_cantidad, width=10)
entry_cantidad.pack(side="left", padx=5)

def agregar_pedido():
    if not tree_med.selection():
        return
    if not entry_cantidad.get().isdigit():
        return

    nombre, _ = tree_med.item(tree_med.selection()[0])["values"]
    pedido.append({"nombre": nombre, "cantidad": int(entry_cantidad.get())})
    entry_cantidad.delete(0, tk.END)
    actualizar_tabla_pedido()

tk.Button(root, text="Agregar", command=agregar_pedido).pack()

# =============================
# PEDIDOS PENDIENTES
# =============================
tk.Label(root, text="Pedidos pendientes / enviados",
         font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=5)

tree_pendientes = ttk.Treeview(root, columns=("folio", "estado"),
                               show="headings", height=6)
tree_pendientes.heading("folio", text="Folio")
tree_pendientes.heading("estado", text="Estado")
tree_pendientes.pack(fill="x", padx=10)

def actualizar_tabla_pendientes():
    tree_pendientes.delete(*tree_pendientes.get_children())
    for p in pedidos_pendientes:
        icono = {
            "Pendiente": "⏸",
            "Enviando": "⏳",
            "Enviado": "✅",
            "Error": "❌"
        }.get(p["estado"], "")
        tree_pendientes.insert("", tk.END, values=(p["folio"], f"{p['estado']} {icono}"))

def refrescar_estados():
    actualizar_tabla_pendientes()
    root.after(1000, refrescar_estados)

# =============================
# FINALIZAR
# =============================
def finalizar_pedido():
    if not pedido:
        return

    ruta, folio = exportar_pedido_excel(pedido)
    info = {"folio": folio, "ruta": ruta, "estado": "Pendiente", "error": None}
    pedidos_pendientes.append(info)

    guardar_pedido_bd(folio, ruta, "Pendiente", pedido)

    threading.Thread(target=enviar_pedido_async, args=(info,), daemon=True).start()

    pedido.clear()
    actualizar_tabla_pedido()

tk.Button(root, text="Finalizar pedido", bg="green",
          fg="white", command=finalizar_pedido).pack(pady=10)

# =============================
# REINTENTAR
# =============================
def reenviar_pedido():
    if not tree_pendientes.selection():
        return
    index = tree_pendientes.index(tree_pendientes.selection()[0])
    threading.Thread(target=enviar_pedido_async,
                     args=(pedidos_pendientes[index],),
                     daemon=True).start()

tk.Button(root, text="Reintentar envío", command=reenviar_pedido).pack()

# =============================
# INICIAR REFRESCO
# ============================
# 
# 
# 

def envio_automatico():
    for p in pedidos_pendientes:
        if p["estado"] in ("Pendiente", "Error"):
            threading.Thread(
                target=enviar_pedido_async,
                args=(p,),
                daemon=True
            ).start()

    # Reintenta cada 2 minutos (120000 ms)
    root.after(120000, envio_automatico)

refrescar_estados()
envio_automatico()
root.mainloop()
